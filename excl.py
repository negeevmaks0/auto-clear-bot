from openpyxl import load_workbook

import os
import sys



class Excel:
    def __init__(self):
        self.dir_ = os.path.dirname(os.path.abspath(__file__))

        self.filepath = os.path.join(self.dir_, 'planung_tasks.xlsx')
        self.filepath_pdf = os.path.join(self.dir_, 'planung_tasks.pdf')

        self.tasks = [
            (range(4, 12), 0, 'A', 8),
            (range(13, 24), 1, 'A', 11),
            (range(3, 15), 2, 'B', 12)
        ]


    def update_data(self, data, monat_):
        wb = load_workbook(self.filepath)
        ws = wb.active

        ws['A1'] = monat_

        for r_range, idx, col, length in self.tasks:
            data[idx].extend(['-----'] * (length - len(data[idx])))

            for row, val in zip(r_range, data[idx]):
                ws[f'{col}{row}'] = val

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        wb.save(self.filepath)
        wb.close()

        return self.convert_to_pdf()


    def convert_to_pdf(self):
        if sys.platform.startswith('win'):
            return self._convert_windows()

        else:
            return self._convert_linux()


    def _convert_windows(self):
        import win32com.client

        abs_excel_path = os.path.abspath(self.filepath)
        abs_pdf_path = os.path.abspath(self.filepath_pdf)

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False

        wb = excel.Workbooks.Open(abs_excel_path)
        wb.ActiveSheet.ExportAsFixedFormat(0, abs_pdf_path, Quality=0)
        wb.Close(False)

        excel.Quit()

        return self.filepath_pdf


    def _convert_linux(self):
        import subprocess

        command = [
            'libreoffice',
            '--headless',
            '--convert-to',
            'pdf',
            self.filepath,
            '--outdir',
            self.dir_
        ]

        subprocess.run(command, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        return self.filepath_pdf
